#!/usr/bin/env python3
"""
Excel Table Extractor - Extract specific ranges to new workbooks and render as PNG

This approach is much more reliable than PDF cropping:
1. Extract only the target range from the source Excel
2. Create a new workbook containing just that range (with all formatting)
3. Render the new workbook - the entire image IS the table
4. No cropping needed!
"""

import json
import sys
import base64
import os
import tempfile
import subprocess
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.utils import range_boundaries
from copy import copy


def extract_and_render(excel_base64: str, sheet_name: str, cell_range: str) -> dict:
    """
    Extract a range from Excel, create new workbook with just that range,
    convert to PNG via LibreOffice.
    
    Args:
        excel_base64: Base64 encoded Excel file
        sheet_name: Target sheet name (e.g., "S&U ")
        cell_range: Target range (e.g., "A6:N27")
    
    Returns:
        dict: {"success": bool, "image": str, "error": str}
    """
    source_path = None
    output_xlsx = None
    output_dir = None
    
    try:
        # Decode input Excel to temporary file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            f.write(base64.b64decode(excel_base64))
            source_path = f.name
        
        # Load source workbook
        src_wb = load_workbook(source_path, data_only=False)
        
        # Find target sheet (case-insensitive, partial match)
        target_sheet = None
        for sheet in src_wb.worksheets:
            if sheet.title.strip().lower() == sheet_name.strip().lower():
                target_sheet = sheet
                break
        
        if not target_sheet:
            # Try partial match
            for sheet in src_wb.worksheets:
                if (sheet_name.strip().lower() in sheet.title.strip().lower() or 
                    sheet.title.strip().lower() in sheet_name.strip().lower()):
                    target_sheet = sheet
                    break
        
        if not target_sheet:
            return {
                "success": False,
                "error": f"Sheet '{sheet_name}' not found. Available sheets: {[s.title for s in src_wb.worksheets]}"
            }
        
        print(f"Found target sheet: '{target_sheet.title}'", file=sys.stderr)
        
        # Parse range "A6:N27"
        try:
            min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        except Exception as e:
            return {
                "success": False,
                "error": f"Invalid range '{cell_range}': {str(e)}"
            }
        
        print(f"Extracting range {cell_range}: cols {min_col}-{max_col}, rows {min_row}-{max_row}", file=sys.stderr)
        
        # Create new workbook with just this range
        new_wb = Workbook()
        new_ws = new_wb.active
        new_ws.title = "Table"
        
        # Copy cells with ALL formatting
        for src_row_idx, src_row in enumerate(range(min_row, max_row + 1), 1):
            for src_col_idx, src_col in enumerate(range(min_col, max_col + 1), 1):
                src_cell = target_sheet.cell(row=src_row, column=src_col)
                new_cell = new_ws.cell(row=src_row_idx, column=src_col_idx)
                
                # Copy value (handle formulas - get calculated value if possible)
                if src_cell.value is not None:
                    new_cell.value = src_cell.value
                
                # Copy ALL formatting
                if src_cell.has_style:
                    new_cell.font = copy(src_cell.font)
                    new_cell.fill = copy(src_cell.fill)
                    new_cell.border = copy(src_cell.border)
                    new_cell.alignment = copy(src_cell.alignment)
                    new_cell.number_format = src_cell.number_format
                    new_cell.protection = copy(src_cell.protection)
        
        # Copy column widths
        for src_col in range(min_col, max_col + 1):
            src_col_letter = target_sheet.cell(row=1, column=src_col).column_letter
            new_col_idx = src_col - min_col + 1
            new_col_letter = new_ws.cell(row=1, column=new_col_idx).column_letter
            
            if src_col_letter in target_sheet.column_dimensions:
                src_width = target_sheet.column_dimensions[src_col_letter].width
                if src_width:
                    new_ws.column_dimensions[new_col_letter].width = src_width
        
        # Copy row heights
        for src_row in range(min_row, max_row + 1):
            new_row_idx = src_row - min_row + 1
            if src_row in target_sheet.row_dimensions:
                src_height = target_sheet.row_dimensions[src_row].height
                if src_height:
                    new_ws.row_dimensions[new_row_idx].height = src_height
        
        # Copy merged cells (adjusted to new positions)
        for merged_range in target_sheet.merged_cells.ranges:
            # Check if merge overlaps with our target range
            if (merged_range.min_row <= max_row and merged_range.max_row >= min_row and
                merged_range.min_col <= max_col and merged_range.max_col >= min_col):
                
                # Adjust merge coordinates to new workbook
                new_min_row = max(1, merged_range.min_row - min_row + 1)
                new_max_row = min(max_row - min_row + 1, merged_range.max_row - min_row + 1)
                new_min_col = max(1, merged_range.min_col - min_col + 1)
                new_max_col = min(max_col - min_col + 1, merged_range.max_col - min_col + 1)
                
                if new_min_row <= new_max_row and new_min_col <= new_max_col:
                    new_ws.merge_cells(
                        start_row=new_min_row, start_column=new_min_col,
                        end_row=new_max_row, end_column=new_max_col
                    )
        
        # Save new workbook
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            output_xlsx = f.name
        new_wb.save(output_xlsx)
        
        print(f"Created extracted workbook: {output_xlsx}", file=sys.stderr)
        
        # Convert to PDF with LibreOffice
        output_dir = tempfile.mkdtemp()
        pdf_result = subprocess.run([
            'soffice', '--headless', '--convert-to', 'pdf',
            '--outdir', output_dir, output_xlsx
        ], capture_output=True, text=True, timeout=30)
        
        if pdf_result.returncode != 0:
            return {
                "success": False,
                "error": f"LibreOffice PDF conversion failed: {pdf_result.stderr}"
            }
        
        pdf_path = os.path.join(output_dir, os.path.basename(output_xlsx).replace('.xlsx', '.pdf'))
        if not os.path.exists(pdf_path):
            return {
                "success": False,
                "error": "LibreOffice did not generate PDF file"
            }
        
        print(f"Generated PDF: {pdf_path}", file=sys.stderr)
        
        # Convert PDF to PNG with poppler
        png_base = os.path.join(output_dir, 'output')
        png_result = subprocess.run([
            'pdftoppm', '-png', '-r', '150', '-singlefile',
            pdf_path, png_base
        ], capture_output=True, text=True, timeout=15)
        
        if png_result.returncode != 0:
            return {
                "success": False,
                "error": f"PDF to PNG conversion failed: {png_result.stderr}"
            }
        
        png_path = png_base + '.png'
        if not os.path.exists(png_path):
            return {
                "success": False,
                "error": "PNG file was not generated"
            }
        
        print(f"Generated PNG: {png_path}", file=sys.stderr)
        
        # Read and return as base64
        with open(png_path, 'rb') as f:
            image_data = base64.b64encode(f.read()).decode('utf-8')
        
        return {
            "success": True,
            "image": image_data,
            "method": "python-openpyxl-extraction",
            "extracted_range": cell_range,
            "source_sheet": target_sheet.title
        }
        
    except Exception as e:
        return {
            "success": False,
            "error": f"Extraction failed: {str(e)}"
        }
    
    finally:
        # Cleanup temp files
        for path in [source_path, output_xlsx]:
            if path and os.path.exists(path):
                try:
                    os.unlink(path)
                except:
                    pass
        
        if output_dir and os.path.exists(output_dir):
            try:
                import shutil
                shutil.rmtree(output_dir)
            except:
                pass


def main():
    """
    Main entry point - read JSON from stdin, process, write JSON to stdout
    """
    try:
        # Read input from stdin
        input_data = json.loads(sys.stdin.read())
        
        excel_base64 = input_data['excelBase64']
        sheet_name = input_data['sheetName']
        cell_range = input_data['cellRange']
        
        # Process
        result = extract_and_render(excel_base64, sheet_name, cell_range)
        
        # Output result as JSON
        print(json.dumps(result))
        
    except Exception as e:
        # Error response
        error_result = {
            "success": False,
            "error": f"Script error: {str(e)}"
        }
        print(json.dumps(error_result))
        sys.exit(1)


if __name__ == "__main__":
    main()